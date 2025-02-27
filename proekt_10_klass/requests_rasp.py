import urllib.request
import requests
import urllib
import openpyxl
import os
import asyncio

class ClassOfGetRasp:
    def id_file_of_rasp(self):
        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'}

            response = requests.get(r"""https://sites.google.com/view/perspectiva99/%D0%B3%D0%BB%D0%B0%D0%B2%D0%BD%D0%B0%D1%8F-%D1%81%D1%82%D1%80%D0%B0%D0%BD%D0%B8%D1%86%D0%B0""", headers=headers, timeout=31)
        except:
            return False

        response = response.text
        response = response

        index = response.find("data-embed-doc-id")
        count_files = response.count("data-embed-doc-id")

        if count_files > 1:
            index = response.find("data-embed-doc-id", index + 1)

        response = response[index:index+100]
        response = response[response.find('"') + 1:index+100]
        result = response[:response.find('"')]
        return result
    
    def download_rasp(self, name):
        try:
            file_xlsx = urllib.request.urlopen(f"https://docs.google.com/spreadsheets/d/{self.id_file_of_rasp()}/export?format=xlsx").read()

            # file_xlsx = urllib.request.urlopen(f"https://drive.usercontent.google.com/u/0/uc?id={id_file_of_rasp()}&export=download").read()
            
            with open(file=f'{name}.xlsx', mode='wb') as file:
                file.write(file_xlsx)
                file.close()

            ################################################################################## работает не трогай 
            file = f'{name}.xlsx'
            new_file = f'{name}.xlsx'

            # Открываем существующий файл Excel:
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active

            # Создаём список объединенных ячеек:
            merged_cells = list(sheet.merged_cells.ranges)

            # Проходим по всем объединённым ячейкам:
            for merged_cell in merged_cells:
                # Разъединяем ячейки:
                sheet.unmerge_cells(str(merged_cell))
                # Копируем значение в каждую ячейку:
                top_left_cell = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col)
                for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                    for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                        sheet.cell(row=row, column=col).value = top_left_cell.value

            # Сохраняем изменения:
            workbook.save(new_file)
            ##################################################################################
            return True
        except:
            return False
        
    async def check_files(self):
        rasp = openpyxl.open('расписание.xlsx').active.iter_cols(values_only=True)

        self.download_rasp('расписание0')
        try:
            await asyncio.sleep(10)
            rasp0 = openpyxl.open('расписание0.xlsx').active.iter_cols(values_only=True)

            res = []
            for i in rasp:
                res.append(i)

            res0 = []
            for i in rasp0:
                res0.append(i)

            if not (res == res0):
                os.remove('расписание.xlsx')
                os.rename('расписание0.xlsx', 'расписание.xlsx')

            else:
                os.remove('расписание0.xlsx')

            return not (res == res0)
        except:
            print('ошибка в check_files')

    def return_rasp_for_user(self, grade):
        try:
            grade = float(grade)
        except:
            None

        xlsx = openpyxl.open('расписание.xlsx', data_only=True).active

        all_cols = list(xlsx.iter_cols(values_only=True))
        all_rows = list(xlsx.iter_rows(values_only=True))

        col_of_time = all_cols[1]

        col = ()

        for i in all_cols:
            if grade in i:
                col = i

        max_lessons = 0

        try:
            for i in all_cols[0][col.index(grade)+1:]:
                try:
                    max_lessons = int(i)
                except:
                    break
        except:
            return False
        
        index_grade_in_cols = col.index(grade)

        col = list(col[index_grade_in_cols+1:index_grade_in_cols+max_lessons+1])
        col_of_time = list(col_of_time[index_grade_in_cols+1:index_grade_in_cols+max_lessons+1])

        for i in col[::-1]:
            if i == None:
                col.pop(len(col)-1)
            else:
                break

        for i in range(0, len(col)-1):
            if col[i] == None:
                col[i] = 'Окно'

        for i in range(0, len(col)):
            col[i] = f'{i+1}) <i>{col_of_time[i]}</i> - <strong>{col[i]}</strong>'

        res_lessons = ''

        for i in col:
            res_lessons += f'{i}\n'

        try:
            grade = str(int(grade))
        except:
            None
            
        try:
            date_time = str(all_rows[0][0].date())
        except:
            date_time = str(all_rows[0][0])
        
        return date_time + '\n\n' + grade + '\n' + res_lessons + '\n' + r'<a href="https://sites.google.com/view/perspectiva99/%D0%B3%D0%BB%D0%B0%D0%B2%D0%BD%D0%B0%D1%8F-%D1%81%D1%82%D1%80%D0%B0%D0%BD%D0%B8%D1%86%D0%B0">Ссылка на расписание</a>' 

    def write_all_classes(self):
        xlsx = openpyxl.open('расписание.xlsx').active

        all_classes = []

        all_rows = list(xlsx.iter_rows(values_only=True))

        for row in all_rows:
            if row[0] == '#':
                for i in row[1:]:
                    if i != None:
                        all_classes.append(i)

        with open(file = 'all_classes.txt', mode='w+', encoding='utf-8') as file:
            for i in all_classes:
                try:
                    i = int(i)
                except:
                    None

                file.write(f'{i}\n')


    def return_all_classes(self):
        with open(file = 'all_classes.txt', mode='r', encoding='utf-8') as file:
            all_classes = file.readlines()

        for i in range(0, len(all_classes)):
            all_classes[i] = all_classes[i][:all_classes[i].index('\n')]

        for i in range(0, len(all_classes)):
            try:
                all_classes[i] = int(all_classes[i])
            except:
                None
        
        return all_classes